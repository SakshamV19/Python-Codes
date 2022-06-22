import openpyxl 
path = "D:\\Excel1.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
mrow = sheet_obj.max_row
mcol = sheet_obj.max_column
for row in sheet_obj.iter_rows(min_row = 1, min_col =1, max_row= mrow, max_col = mcol):
    for cell in row:
        print(cell.value, end =" ")
    print()
    
        

